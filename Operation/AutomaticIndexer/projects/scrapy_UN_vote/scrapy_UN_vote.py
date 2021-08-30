# coding:utf-8

"""
# 自动化操作：获取联合国投票信息

模式1：
1. 自动操作联合国网站的各年投票结果并保存：
1.10. 自动爬取各年的投票结果选项并自动搜索；
1.20. 对于每一年投票结果，爬取其中的html table类型的表格；
1.30. 表格转换成易于操作的电子表格格式，如excel格式；
1.40. 对立面的数据进行梳理；

模式2：TODO
1. 自动操作联合国网站的各年投票结果并保存：
1.10. 模拟器选中某一年的投票结果选项并自动搜索；
1.20. 对于每一年投票结果，爬取其中的html table类型的表格；
1.30. 表格转换成易于操作的电子表格格式，如excel格式；
1.40. 对立面的数据进行梳理；

"""

import sys
import numpy
import math
import re
import time

import pandas
from pandas import DataFrame
from selenium import webdriver as Webdriver
from lxml import etree
# from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
# from selenium.webdriver.support import expected_conditions as EC
# from selenium.webdriver.support.wait import WebDriverWait
import requests
import scrapy
from pyquery import PyQuery
import json
import openpyxl


#----------- choose the mode ------------- #


"""
模式'1'：
模式'2'：

"""
args_mode = '1'

# -------------------------------------
# ------------ set parameters -------------------
args_web_url_root=r'https://digitallibrary.un.org'
args_web_url_full_initPage= args_web_url_root + '/search?ln=zh_CN&cc=Voting+Data&p=&f=&rm=&ln=zh_CN&sf=&so=d&rg=50&c=Voting+Data&c=&of=hb&fti=0&fct__2=General+Assembly&fct__9=Vote&fct__3=2013&fti=0&fct__2=General+Assembly&fct__9=Vote'
args_web_url_full_nextPage = r'https://digitallibrary.un.org/search?ln=zh_CN&jrec=1&fct__3=2013&fct__2=General+Assembly&fct__2=General+Assembly&cc=Voting+Data&fct__9=Vote&fct__9=Vote'
args_path_data_output = r'./AutomaticIndexer/data/'
args_year_range = [2013,2019]
keyword=''







# -----------------------------------


def main(args_mode):

    switch = {
        '1': function_1(web_url_root=args_web_url_root, web_url_full_initPage=args_web_url_full_initPage, web_url_full_nextPage=args_web_url_full_nextPage, path_data=args_path_data_output, year_range=args_year_range)
        # '2': function_2(web_url=args_web_url_full)
    }
    switch.get(args_mode)
    # function_transferData()
    pass

# 获取输入关键词之后的网页


def function_1(web_url_root, web_url_full_initPage, web_url_full_nextPage, path_data, year_range):
    
    # browser = Webdriver.Chrome()  # 生成浏览器驱动对象
    try:
        
        # 数据结构用之于保存收集的数据
        # data_country_name = ''
        # data_vote_result = ''
        # data_vote = {'country name': data_country_name,
        #              'vote result': data_vote_result}
        # data_meeting_name = ''
        # data_list_vote = [data_vote]
        # data_meeting = {'meeting name': data_meeting_name,
        #                 'vote result': data_list_vote}
        # data_year_name = None
        # data_list_meeting = [data_meeting]
        # data_year = {'year name': data_year_name,
        #                   'meeting': data_list_meeting}
        # data = [data_year]

        ## 设置收集的数据格式
        
        data_list_vote = []
        data = []

        ## 自动爬取联合国网站的各年投票结果并保存：
        list_year_name = list(range(year_range[0],year_range[1],1))
        for year in list_year_name:
            data_list_meeting = []

            # test
            print('year: '+str(year))

            # 更改网址的年份
            pattern_words = re.compile(r'fct__3=\d*[^&]')
            replace_words = r'fct__3='+str(year)
            web_url_full_initPage = re.sub(pattern_words, replace_words, web_url_full_initPage)

            # test
            print(web_url_full_initPage)

            # 提交主页的请求
            request = requests.get(web_url_full_initPage)
            request_text = request.text

            # html转格式为PyQuery格式
            html_mainpage = PyQuery(request_text)

            # # 写入的主页面的html，用于测试时观察
            # with open('request_text.html','a',encoding='utf-8') as file:
            #     file.write(request_text)


            ## 遍历各页
            num_items = int(html_mainpage('.rec-count strong').text()) # 搜索结果个数
            num_pages = math.ceil(num_items/50) # 计算出页数（该网站默认每页50项）
            for page in list(range(num_pages)):

                # data_list_meeting = []

                # test
                print('page: ' + str(page))

                # 收集子链接
                html_items_mainpage = html_mainpage('span.moreinfo:first-child a')

                # 保存用以列表
                list_webURL_meeting=[]
                for item in html_items_mainpage.items():
                    list_webURL_meeting.append(web_url_root + item.attr('href'))
                    pass
                
                ## 收集各决议名称
                html_items_mainpage = html_mainpage('.brief-options')
                text_items_mainpage = html_items_mainpage.text()

                # 提取决议名称号
                pattern_words = re.compile(r'A/RES/\S*')
                list_meeting_names = re.findall(pattern_words,text_items_mainpage)
                # print(html_items_mainpage)

                ## 进入各网页提取数据
                idx_meeting=0 # 本页内，各子链接的访问标记，即for循环了第几次。
                for weburl_meeting in list_webURL_meeting[:]:
                    # test
                    print('meeting name: ' + list_meeting_names[list_webURL_meeting.index(weburl_meeting)])
                    # print(weburl_meeting)

                    # 提交子链接请求
                    request = requests.get(weburl_meeting)
                    request_text = request.text

                    # html转格式为PyQuery格式
                    html_meetingpage = PyQuery(request_text)

                    ## 1.20. 对于每一年投票结果，爬取其中的html table类型的文本数据；
                    html_items_meeting = html_meetingpage('#details-collapse > div:nth-last-child(2) > span.value.col-xs-12.col-sm-9.col-md-10')
                    text_vote_result = html_items_meeting.text().split('\n')

                    ## 1.30. 分离投票结果和国家于文本中
                    list_country = []
                    list_vote_result = []
                    pattern_words = re.compile('[YNA][ ]')
                    for idx in list(range(len(text_vote_result))):                        
                        if re.search(pattern_words, text_vote_result[idx]) == None: # 判断是否匹配
                            list_vote_result.append('O')
                            list_country.append(text_vote_result[idx])
                        else:
                            list_vote_result.append(re.search(pattern_words,text_vote_result[idx]).group()[0])
                            list_country.append(re.sub(pattern_words,'',text_vote_result[idx]))
                            pass
                        pass
                        
                    
                    ## 1.40. 提取的数据转换成初始数据结构，以便于后续用json存储；
                    data_meeting_name = list_meeting_names[idx_meeting]
                    data_list_vote = []
                    for data_country_name, data_vote_result in zip(list_country,list_vote_result):
                        data_vote = {'country name': data_country_name,
                                     'vote result': data_vote_result}
                        data_list_vote.append(data_vote)
                        pass
                    data_meeting = {
                        'meeting name': data_meeting_name, 'vote result': data_list_vote}
                    data_list_meeting.append(data_meeting)

                    idx_meeting = idx_meeting+1

                    pass # for 读取各子链接
                
                ## 更改网址的页数信息，为翻页做准备
                num_item = (page+1)*50+1  # 网址显示的页面的条目的标号

                # 替换为新的页面对应的条目的标号
                pattern_words = re.compile(r'jrec=\d*[^&]')
                replace_words = r'jrec='+str(num_item)
                web_url_full_nextPage = re.sub(pattern_words,replace_words,web_url_full_nextPage)

                # 更改网址的年份
                pattern_words = re.compile(r'fct__3=\d*[^&]')
                replace_words = r'fct__3=' + str(year)
                web_url_full_nextPage = re.sub(pattern_words, replace_words, web_url_full_nextPage)

                # test
                print(web_url_full_nextPage)

                # 提交翻页的请求
                request = requests.get(web_url_full_nextPage)
                request_text = request.text
                # html转格式为PyQuery格式
                html_mainpage = PyQuery(request_text)

                pass # for 翻页
            
            ## 1.40. 提取的数据转换成初始数据结构；
            data_year_name = str(year)
            data_year = {
                'year name': data_year_name, 'meeting': data_list_meeting}
            data.append(data_year)
        

            pass # for 更换年份主页面


        ## 保存提取的数据为json格式
        with open('data_init_1.json', 'w') as f:
            json.dump(data, f)





    finally:
        # browser.close()
        pass

    pass # try 抛出异常

def function_transferData(path_data):
    ## 读取存储的json数据文件
    with open(path_data+'data_init.json') as f:
        data = json.load(f)
    

    ### ---- 模式1：分开所有年份 ---
    ## 1.50. 对提取的数据转换；
    # excel_workbook = openpyxl.Workbook() # 创建excel数据簿

    # 遍历各年
    for idx_year in list(range(len(data))): 
        # 对于本年，生成各次会议的生成各数据的df结构
        keys_country = list()
        index_meetings = []
        df_eachVoteResult = []
        for idx_meeting in list(range(len(data[idx_year]['meeting']))):
            # df_eachVoteResult.append(DataFrame(data = [x['vote result'] for x in data[idx_year]['meeting'][idx_meeting]]))
            df_eachVoteResult.append(DataFrame(data = data[idx_year]['meeting'][idx_meeting]['vote result']))
            # 求并集，对于各个会议的各国序列
            keys_country = list(set(keys_country) | set([x['country name'] for x in data[idx_year]['meeting'][idx_meeting]['vote result']]))
            # 提取所有的会议名称
            index_meetings.append(data[idx_year]['meeting'][idx_meeting]['meeting name'])
            pass
        # 合并生成一个大表，该大表行是本年度各会议，列是本年度涉及的所有投票的国家，内容是投票结果。
        df_voteResult = DataFrame(columns=(keys_country),index=(index_meetings))
        for meeting in index_meetings: # 遍历本年度的各会议
            idx_meeting = index_meetings.index(meeting) # 获取该会议所在索引
            for country in list(df_eachVoteResult[idx_meeting].loc[:,'country name']): # 遍历该会议的国家
                idx_country = df_eachVoteResult[idx_meeting][(df_eachVoteResult[idx_meeting]['country name'] == country)].index.tolist()[0] # 获取该国家所在索引
                df_voteResult.loc[meeting, country] = df_eachVoteResult[idx_meeting].loc[idx_country,'vote result'] # 把该会议的该国家的投票结果赋值给大表
                pass
            pass
        ## 1.55. 写出本年文件到数据表excel的各sheet
        # # pandas自带的写出代码块，不能把各数据表写入一个Excel工作簿
        # excel_writer=pandas.ExcelWriter(path_data+'data_output_'+data[idx_year]['year name']+'.xlsx')
        # df_voteResult.to_excel(excel_writer = excel_writer,sheet_name=data[idx_year]['year name'])
        # excel_writer.save()

        # pandas调用的写出代码块，可以把各数据表写入一个Excel工作簿
        # book = openpyxl.load_workbook(path_data+'data_output_'+data[idx_year]['year name']+'.xlsx')
        book = openpyxl.load_workbook(path_data+'data_output_eachyear.xlsx')
        writer = pandas.ExcelWriter(path_data+'data_output_eachyear.xlsx', engine='openpyxl')
        writer.book = book
        df_voteResult.to_excel(writer, sheet_name=data[idx_year]['year name'])
        writer.save()
        pass # for 遍历年份




    ### ---- 模式2：合并所有年份 ---

    # 遍历各年
    # 生成所有年各次会议的生成各数据的df结构
    keys_country = list()
    index_meetings = []
    df_eachVoteResult = []
    for idx_year in list(range(len(data))): 
        # 遍历本年会议
        for idx_meeting in list(range(len(data[idx_year]['meeting']))):
            df_eachVoteResult.append(DataFrame(data = data[idx_year]['meeting'][idx_meeting]['vote result']))
            # 求并集，对于各个会议的各国序列
            keys_country = list(set(keys_country) | set([x['country name'] for x in data[idx_year]['meeting'][idx_meeting]['vote result']]))
            # 提取所有的会议名称
            index_meetings.append(data[idx_year]['meeting'][idx_meeting]['meeting name'])
            pass # for 遍历本年会议

        pass # for 遍历年份
        
    # 合并生成一个大表，该大表行是本年度各会议，列是本年度涉及的所有投票的国家，内容是投票结果。
    df_voteResult = DataFrame(columns=(keys_country),index=(index_meetings))
    for meeting in index_meetings: # 遍历所有年的各会议
        idx_meeting = index_meetings.index(meeting) # 获取所有年的会议所在索引
        for country in list(df_eachVoteResult[idx_meeting].loc[:,'country name']): # 遍历该会议的国家
            idx_country = df_eachVoteResult[idx_meeting][(df_eachVoteResult[idx_meeting]['country name'] == country)].index.tolist()[0] # 获取该国家所在索引
            df_voteResult.loc[meeting, country] = df_eachVoteResult[idx_meeting].loc[idx_country,'vote result'] # 把该会议的该国家的投票结果赋值给大表
            pass
        pass

    ## 1.55. 写出本年文件到数据表excel的各sheet

    # # pandas自带的写出代码块，不能把各数据表写入一个Excel工作簿
    # excel_writer=pandas.ExcelWriter(path_data+'data_output_'+data[idx_year]['year name']+'.xlsx')
    # df_voteResult.to_excel(excel_writer = excel_writer,sheet_name=data[idx_year]['year name'])
    # excel_writer.save()

    # pandas调用的写出代码块，可以把各数据表写入一个Excel工作簿
    # book = openpyxl.load_workbook(path_data+'data_output_'+data[idx_year]['year name']+'.xlsx')
    book = openpyxl.load_workbook(path_data+'data_output_allyear.xlsx')
    writer = pandas.ExcelWriter(path_data+'data_output_allyear.xlsx', engine='openpyxl')
    writer.book = book
    df_voteResult.to_excel(writer, sheet_name=data[0]['year name']+'-'+data[-1]['year name'])
    writer.save()





# TODO ?为什么会连带运行下面的函数？
# def function_2(web_url):
#     # 获取输入关键词之后的网页
#     browser = Webdriver.Chrome()  # 生成浏览器驱动对象
#     try:
#         ## 假设前面已经进行过高级搜索和输入关键词并点击搜索按钮了。
#         # 获取并载入该网页
#         browser.get(web_url)
#
#         # 选择每页显示50项
#         web_url = browser.current_url  # 获取点击之后的网页网址
#         button_handle = browser.find_element_by_class_name(
#             'select2-selection__rendered')  # 下拉菜单：每页显示更多项
#         button_handle.click()  # 模拟点击
#         web_url = browser.current_url  # 获取点击之后的网页网址
#         time.sleep(1)  # 等待
#
#         # 模式匹配寻找每页显示50项的那个网页元素
#         # 弃用以下注释部分
#         # html=requests.get(web_url)
#         # html.raise_for_status()
#         # html.encoding = html.apparent_encoding  # 判断是否爬取成功
#         # html_text=html.text
#         # html_tree=etree.HTML(html_text)
#         html = browser.find_element_by_xpath("//*").get_attribute("outerHTML")
#         html_tree = etree.HTML(html)
#         html_node = html_tree.xpath(
#             "//ul[@class='select2-results__options']/li[text()=' 每页 50 条 ']")
#         # pattern_words = r'(.+50)'
#         # urlre = re.compile(pattern_words)
#         # element_text = urlre.findall(html_tree)
#         button_handle = browser.find_element_by_id(
#             html_node[0].attrib['id'])  # 选择每页显示50项
#         button_handle.click()  # 模拟点击
#         time.sleep(3)  # 等待
#
#         # 每页选择全部项目添加到标记结果列表然后翻页，直至最后一页。
#         web_url = browser.current_url
#         html_element = browser.find_element_by_id('pageCount.top')
#         page_last = int(html_element.text)
#         pages = range(1, page_last+1-1)
#         for page in pages:
#             button_handle = browser.find_element_by_name(
#                 'SelectPage')  # “选择页面”按钮
#             button_handle.click()  # 模拟点击
#             time.sleep(1)  # 等待
#             button_handle = browser.find_element_by_class_name(
#                 'addToMarkedListButton')  # 添加到标记结果列表
#             button_handle.click()  # 模拟点击
#             time.sleep(1)  # 等待
#             button_handle = browser.find_element_by_class_name(
#                 'paginationNext.snowplow-navigation-nextpage-bottom')  # 下一页
#             button_handle.click()  # 模拟点击
#             time.sleep(1)  # 等待
#             web_url = browser.current_url  # 获取执行了下一页之后的网页网址
#             pass
#         button_handle = browser.find_element_by_name('SelectPage')  # “选择页面”按钮
#         button_handle.click()  # 模拟点击
#         time.sleep(1)  # 等待
#         button_handle = browser.find_element_by_class_name(
#             'addToMarkedListButton')  # 添加到标记结果列表
#         button_handle.click()  # 模拟点击
#         time.sleep(1)  # 等待
#         web_url = browser.current_url  # 获取执行了下一页之后的网页网址
#
#     finally:
#         # browser.close()
#         pass


if __name__ == '__main__':
    sys.exit(main(args_mode))
